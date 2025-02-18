# report_generator.py
import pandas as pd
import streamlit as st
from typing import Dict, Optional
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment
from prompt_list import (DOCUMENT_INFO_PROMPTS, EMPLOYER_INFO_PROMPTS, PLAN_ADMIN_PROMPTS,
                        CONTRIBUTION_TYPES_PROMPTS, SERVICE_TYPES_PROMPTS,
                        ELIGIBILITY_TYPES_PROMPTS, VESTING_TYPES_PROMPTS,
                        COMPENSATION_TYPES_PROMPTS,
                        CONTRIBUTION_ALLOCATION_TYPES_PROMPTS,
                        SAFE_HARBOR_PROMPTS,
                        MATCHING_CONTRIBUTION_PROMPTS, NON_ELECTIVE_PROMPTS,
                        PREVAILING_WAGE_PROMPTS, GENERAL_PLAN_PROMPTS,
                        RETIREMENT_AND_DISBALITIY_PROMPTS,
                        PERMITTED_ELECTION_PROMPTS, DISTRIBUTION_PROMPTS,
                        PARTICIPATING_EMPLOYER_PROMPTS, SUBSEQUENT_CYCLE4_EMPLOYER_PROMPTS,
                        OTHER_PROVISIONS_PROMPTS, DOCUMENT_REQUESTS_PROMPTS,
                        SUPPORTING_FORMS_INFORMATION_PROMPTS, ADMINISTRATIVE_FORMS_PROMPTS,
                        GENERAL_PROMPTS)
from datetime import datetime
from threading import Lock
import logging

class ReportGenerationTracker:
    def __init__(self):
        self.reports: Dict[str, Dict] = {}
        self.lock = Lock()
    
    def start_generation(self, user_id: str):
        with self.lock:
            self.reports[user_id] = {
                'start_time': datetime.now(),
                'status': 'generating'
            }
    
    def complete_generation(self, user_id: str):
        with self.lock:
            if user_id in self.reports:
                self.reports[user_id].update({
                    'status': 'completed',
                    'end_time': datetime.now()
                })

    def fail_generation(self, user_id: str, error: str):
        with self.lock:
            if user_id in self.reports:
                self.reports[user_id].update({
                    'status': 'failed',
                    'error': error,
                    'end_time': datetime.now()
                })

class ConflictReportGenerator:
    def __init__(self, previous_plan: Dict, restated_plan: Dict, user_id: Optional[str] = None):
        self.previous_plan = previous_plan
        self.restated_plan = restated_plan
        self.user_id = user_id or st.session_state.get('user_id', 'default')
        self.tracker = ReportGenerationTracker()
        
        self.all_prompts = {
            'document_info': DOCUMENT_INFO_PROMPTS,
            'employer_info': EMPLOYER_INFO_PROMPTS,
            'plan_admin': PLAN_ADMIN_PROMPTS,
            'contribution_types': CONTRIBUTION_TYPES_PROMPTS,
            'service_types': SERVICE_TYPES_PROMPTS,
            'eligibility_types': ELIGIBILITY_TYPES_PROMPTS,
            'vesting_types': VESTING_TYPES_PROMPTS,
            'compensation_types': COMPENSATION_TYPES_PROMPTS,
            'contribution_allocation_types': CONTRIBUTION_ALLOCATION_TYPES_PROMPTS,
            'safe_harbor_types': SAFE_HARBOR_PROMPTS,
            'matching_contribution_types': MATCHING_CONTRIBUTION_PROMPTS,
            'nonelective_contribution_types': NON_ELECTIVE_PROMPTS,
            'prevailing_wage_contribution_types': PREVAILING_WAGE_PROMPTS,
            'general_plan_provision_contribution_types': GENERAL_PLAN_PROMPTS,
            'retirement_and_distribution_contribution_types': RETIREMENT_AND_DISBALITIY_PROMPTS,
            'distribution_contribution_types': DISTRIBUTION_PROMPTS,
            'permitted_elections_types': PERMITTED_ELECTION_PROMPTS,
            'participating_employer_types': PARTICIPATING_EMPLOYER_PROMPTS,
            'subsequent_employer_types': SUBSEQUENT_CYCLE4_EMPLOYER_PROMPTS,
            'other_provision_employer_types': OTHER_PROVISIONS_PROMPTS,
            'document_request_employer_types': DOCUMENT_REQUESTS_PROMPTS,
            'supporting_forms_information_types': SUPPORTING_FORMS_INFORMATION_PROMPTS,
            'administrative_forms_types': ADMINISTRATIVE_FORMS_PROMPTS,
            'general_types': GENERAL_PROMPTS
        }

    def get_all_attributes(self) -> list:
        """Get all attributes from all prompt types"""
        all_attributes = []
        for section, prompts in self.all_prompts.items():
            for attr in prompts.keys():
                all_attributes.append((section, attr))
        return all_attributes

    def flatten_results(self, results: Dict) -> Dict:
        """Flatten nested dictionary of results with section prefix"""
        flattened = {}
        for section, section_data in results.items():
            if section_data:
                for attr, value in section_data.items():
                    flattened[f"{section}_{attr}"] = value
        return flattened

    def generate_excel_report(self) -> BytesIO:
        """Generate Excel report with conflicts analysis"""
        try:
            # Track generation start
            self.tracker.start_generation(self.user_id)
            
            rows = []
            si_no = 1

            # Add metadata
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            rows.append({
                'SI No': '',
                'Attributes': 'Report Generated',
                'Previous Plan Document': timestamp,
                'Restated Plan Document': f'User ID: {self.user_id}',
                'Conflict': ''
            })

            # Process each section and its attributes
            for section, prompts in self.all_prompts.items():
                for attr in prompts.keys():
                    prev_value = (self.previous_plan.get(section, {}) or {}).get(attr, 'NA')
                    rest_value = (self.restated_plan.get(section, {}) or {}).get(attr, 'NA')
                    
                    conflict = 'YES' if prev_value != rest_value and prev_value != 'NA' and rest_value != 'NA' else 'NO'
                    display_attr = f"{section.replace('_', ' ').title()} - {attr.replace('_', ' ').title()}"
                    
                    rows.append({
                        'SI No': si_no,
                        'Attributes': display_attr,
                        'Previous Plan Document': prev_value,
                        'Restated Plan Document': rest_value,
                        'Conflict': conflict
                    })
                    si_no += 1

            df = pd.DataFrame(rows)
            
            # Create Excel file in memory
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Conflicts Report')
                self._apply_excel_styling(writer, df)

            self.tracker.complete_generation(self.user_id)
            output.seek(0)
            return output
            
        except Exception as e:
            error_msg = f"Error generating report for user {self.user_id}: {str(e)}"
            self.tracker.fail_generation(self.user_id, str(e))
            logging.error(error_msg)
            raise Exception(error_msg)

    def _apply_excel_styling(self, writer, df):
        """Apply styling to Excel worksheet"""
        worksheet = writer.sheets['Conflicts Report']
        
        # Style the header row
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Auto-adjust columns' width
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(col)
            ) + 2
            worksheet.column_dimensions[chr(65 + idx)].width = max_length
            
        # Center align SI No and Conflict columns
        for row in worksheet.iter_rows(min_row=2):
            row[0].alignment = Alignment(horizontal='center')  # SI No
            row[4].alignment = Alignment(horizontal='center')  # Conflict
            
            # Color code conflict cells
            if row[4].value == 'YES':
                row[4].fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
            elif row[4].value == 'NO':
                row[4].fill = PatternFill(start_color="D9FFD9", end_color="D9FFD9", fill_type="solid")

def download_excel_report(previous_plan: Dict, restated_plan: Dict, user_id: Optional[str] = None):
    """Create and provide download button for Excel report"""
    try:
        # Get user_id from session state if not provided
        user_id = user_id or st.session_state.get('user_id', 'default')
        
        # Create unique key for download button
        button_key = f"download_button_{user_id}"
        
        generator = ConflictReportGenerator(previous_plan, restated_plan, user_id)
        excel_file = generator.generate_excel_report()
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"conflict_report_{user_id}_{timestamp}.xlsx"
        
        st.download_button(
            label="📥 Download Conflict Report",
            data=excel_file,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=button_key
        )
    except Exception as e:
        st.error(f"Error generating Excel report: {str(e)}")